"""
ExcelReportGenerator — produces a two-sheet Excel report using openpyxl.

Sheet1 "摘要"     : header + one summary row
Sheet2 "步骤明细" : all step results, one row per step
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime

# ---------------------------------------------------------------------------
# openpyxl path fix
# The siada_cli_venv ships a broken openpyxl that hard-imports a broken numpy.
# Temporarily remove it from sys.path so the uv-managed openpyxl is used.
# ---------------------------------------------------------------------------
_SIADA_VENV = r"C:\Users\wy_wangbo\.local\share\siada_cli_venv_3.12\Lib\site-packages"
_UV_SITE    = r"C:\Users\wy_wangbo\AppData\Roaming\uv\python\cpython-3.12-windows-x86_64-none\Lib\site-packages"

_removed = [p for p in sys.path if _SIADA_VENV in p]
for _p in _removed:
    sys.path.remove(_p)
if _UV_SITE not in sys.path:
    sys.path.insert(0, _UV_SITE)

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Restore removed paths
for _p in _removed:
    sys.path.append(_p)
# ---------------------------------------------------------------------------

from app.config.settings import REPORTS_DIR
from app.engine.models import TestRecord


# Result foreground colors (dark text on light backgrounds)
_FG_RESULT = {
    "PASS":  "1A8A1A",
    "FAIL":  "CC1A1A",
    "ERROR": "C87020",
    "SKIP":  "888888",
    "ABORT": "C87020",
}
# Row background tints matching HTML report
_ROW_BG = {
    "PASS":  "E8F5E9",
    "FAIL":  "FFEBEE",
    "ERROR": "FFF3E0",
    "SKIP":  "F5F5F5",
    "ABORT": "FFF3E0",
}
_BG_HEADER  = "E8EDF4"
_FG_HEADER  = "222222"
_BG_DATA    = "FFFFFF"
_FG_DATA    = "333333"
_BORDER_CLR = "C8D0DC"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=_FG_DATA, size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Microsoft YaHei")


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _thin_border() -> Border:
    s = Side(style="thin", color=_BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


class ExcelReportGenerator:
    @staticmethod
    def generate(record: TestRecord) -> str:
        """Generate Excel report for *record* and return the file path."""
        os.makedirs(REPORTS_DIR, exist_ok=True)

        ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_plan = re.sub(r'[\\/:*?"<>|]', '_', record.plan_name)
        safe_sn   = record.sn.replace("/", "_").replace("\\", "_")
        filename  = f"{safe_plan}_{safe_sn}_{ts}.xlsx"
        path      = os.path.join(REPORTS_DIR, filename)

        wb = openpyxl.Workbook()
        _build_summary_sheet(wb, record)
        _build_detail_sheet(wb, record)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(path)
        return path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(wb: openpyxl.Workbook, record: TestRecord) -> None:
    ws = wb.create_sheet("摘要")
    ws.sheet_view.showGridLines = False

    headers = ["测试计划", "版本", "运行编号", "开始时间", "结束时间", "总结果",
               "步骤总数", "通过步骤"]
    passed = sum(1 for r in record.step_results if r.result == "PASS")
    total  = len(record.step_results)
    values = [
        record.plan_name, record.plan_version, record.sn,
        record.start_time, record.end_time, record.overall_result,
        total, passed,
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = _fill(_BG_HEADER)
        cell.font      = _font(bold=True, color=_FG_HEADER)
        cell.alignment = _center()
        cell.border    = _thin_border()
        ws.column_dimensions[cell.column_letter].width = max(16, len(h) * 2 + 4)

    result_fg = _FG_RESULT.get(record.overall_result, "333333")
    result_bg = _ROW_BG.get(record.overall_result, "FFFFFF")
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.fill      = _fill(result_bg if col == 6 else _BG_DATA)
        cell.font      = _font(
            color=result_fg if col == 6 else _FG_DATA,
            bold=(col == 6),
        )
        cell.alignment = _center()
        cell.border    = _thin_border()

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22


def _build_detail_sheet(wb: openpyxl.Workbook, record: TestRecord) -> None:
    ws = wb.create_sheet("步骤明细")
    ws.sheet_view.showGridLines = False

    headers    = ["#", "序列", "步骤名称", "结果", "测量值", "单位", "耗时(ms)", "消息"]
    col_widths = [5, 20, 36, 8, 14, 8, 12, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = _fill(_BG_HEADER)
        cell.font      = _font(bold=True, color=_FG_HEADER)
        cell.alignment = _center()
        cell.border    = _thin_border()
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    for row_i, sr in enumerate(record.step_results, 2):
        fg     = _FG_RESULT.get(sr.result, "333333")
        row_bg = _ROW_BG.get(sr.result, "FFFFFF")
        val_str = str(sr.value) if sr.value is not None else "—"
        row_data = [
            row_i - 1, sr.seq_name, sr.step_name, sr.result,
            val_str, sr.unit, sr.duration_ms, sr.message,
        ]
        for col, v in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            cell.fill  = _fill(row_bg)
            cell.font  = _font(
                color=fg if col == 4 else _FG_DATA,
                bold=(col == 4),
            )
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 4, 5, 6, 7) else "left",
                vertical="center",
            )
            cell.border = _thin_border()
        ws.row_dimensions[row_i].height = 20
