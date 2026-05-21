# core/excel_manager.py
import openpyxl
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
import shutil
from openpyxl.utils import column_index_from_string


class PTSExcelManager:
    """PTS Excel测试结果管理类"""

    def __init__(self, template_path: str, output_path: str):
        """
        初始化Excel管理器
        :param template_path: PTS Excel模板文件路径
        :param output_path: 测试结果输出路径
        """
        self.template_path = template_path
        self.output_path = output_path
        self.workbook = None
        self.worksheet = None
        self.test_sheet_name = "测试结果"  # PTS标准sheet名称

        # 初始化Excel文件
        self._init_excel()

    def _init_excel(self):
        """初始化Excel文件（基于模板创建）"""
        try:
            # 情况1: 输出文件已存在且不为空 - 加载现有文件
            if os.path.exists(self.output_path) and os.path.getsize(self.output_path) > 0:
                self.workbook = openpyxl.load_workbook(self.output_path)
                # 确保存在测试结果sheet
                if self.test_sheet_name not in self.workbook.sheetnames:
                    self.workbook.create_sheet(self.test_sheet_name)
                self.worksheet = self.workbook[self.test_sheet_name]

            # 情况2: 模板存在 - 复制模板
            elif os.path.exists(self.template_path):
                # 复制模板文件到输出路径
                output_dir = os.path.dirname(self.output_path)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)

                shutil.copy2(self.template_path, self.output_path)
                self.workbook = openpyxl.load_workbook(self.output_path)

                # 确保存在测试结果sheet
                if self.test_sheet_name not in self.workbook.sheetnames:
                    self.workbook.create_sheet(self.test_sheet_name)
                self.worksheet = self.workbook[self.test_sheet_name]

            # 情况3: 都没有 - 创建新文件
            else:
                self.workbook = openpyxl.Workbook()
                # 重命名默认sheet
                default_sheet = self.workbook.active
                default_sheet.title = self.test_sheet_name
                self.worksheet = self.workbook[self.test_sheet_name]

                # 创建输出目录
                output_dir = os.path.dirname(self.output_path)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir)

            self._init_header()

        except Exception as e:
            raise Exception(f"初始化Excel失败: {str(e)}")

    def _init_header(self):
        """初始化Excel表头（PTS标准格式）- 仅当表头为空时"""
        # 检查是否已有表头（判断第一行第一列是否为空）
        if self.worksheet.max_row == 0 or self.worksheet.cell(row=1, column=1).value is None:
            headers = [
                "测试项编号", "测试项名称", "变量名", "下限值", "上限值",
                "实测值", "测试结果", "测试时间", "备注"
            ]
            # 写入表头
            for col, header in enumerate(headers, 1):
                self.worksheet.cell(row=1, column=col, value=header)
            self.save()

    def get_next_row(self) -> int:
        """获取下一个可写入的行号"""
        # 检查最后一行是否为空（所有列都为空）
        max_row = self.worksheet.max_row
        if max_row == 0:
            return 2  # 第一行是表头

        # 检查最后一行是否为空行
        last_row_empty = True
        for col in range(1, self.worksheet.max_column + 1):
            if self.worksheet.cell(row=max_row, column=col).value is not None:
                last_row_empty = False
                break

        if last_row_empty:
            return max_row
        else:
            return max_row + 1

    def write_test_result(self, test_data: Dict[str, Any]):
        """
        写入单个测试结果
        :param test_data: 测试数据字典，包含：
            - test_id: 测试项编号
            - test_name: 测试项名称
            - var_name: 变量名
            - low_limit: 下限值
            - up_limit: 上限值
            - actual_value: 实测值
            - test_result: 测试结果（Pass/Fail）
            - remarks: 备注（可选）
        """
        try:
            row = self.get_next_row()

            # 写入测试数据（按PTS标准格式）
            self.worksheet.cell(row=row, column=1, value=test_data.get("test_id", row - 1))
            self.worksheet.cell(row=row, column=2, value=test_data.get("test_name", ""))
            self.worksheet.cell(row=row, column=3, value=test_data.get("var_name", ""))
            self.worksheet.cell(row=row, column=4, value=test_data.get("low_limit", ""))
            self.worksheet.cell(row=row, column=5, value=test_data.get("up_limit", ""))
            self.worksheet.cell(row=row, column=6, value=test_data.get("actual_value", ""))
            self.worksheet.cell(row=row, column=7, value=test_data.get("test_result", ""))
            self.worksheet.cell(row=row, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            self.worksheet.cell(row=row, column=9, value=test_data.get("remarks", ""))

            # 保存文件
            self.save()

        except Exception as e:
            raise Exception(f"写入测试结果失败: {str(e)}")

    def write_batch_results(self, results_list: List[Dict[str, Any]]):
        """批量写入测试结果"""
        for result in results_list:
            self.write_test_result(result)

    def read_test_config(self, config_sheet: str = "测试配置") -> List[Dict[str, Any]]:
        """
        从Excel读取测试配置
        :param config_sheet: 配置sheet名称
        :return: 测试配置列表
        """
        try:
            if config_sheet not in self.workbook.sheetnames:
                raise Exception(f"配置sheet {config_sheet} 不存在")

            config_sheet = self.workbook[config_sheet]
            config_list = []

            # 读取表头
            headers = []
            for col in range(1, config_sheet.max_column + 1):
                header = config_sheet.cell(row=1, column=col).value
                if header:
                    headers.append(header)

            # 读取配置数据
            for row in range(2, config_sheet.max_row + 1):
                config_data = {}
                for col, header in enumerate(headers):
                    value = config_sheet.cell(row=row, column=col + 1).value
                    config_data[header] = value

                if config_data:  # 跳过空行
                    config_list.append(config_data)

            return config_list

        except Exception as e:
            raise Exception(f"读取测试配置失败: {str(e)}")

    def save(self):
        """保存Excel文件"""
        try:
            # 确保输出目录存在
            output_dir = os.path.dirname(self.output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            self.workbook.save(self.output_path)
        except Exception as e:
            raise Exception(f"保存Excel失败: {str(e)}")

    def clear_test_results(self, confirm: bool = True):
        """
        清空测试结果（保留表头）
        :param confirm: 是否需要确认
        """
        if confirm:
            response = input("确认要清空所有测试结果吗？(y/n): ")
            if response.lower() != 'y':
                print("操作取消")
                return

        # 删除所有数据行（保留表头）
        for row in range(self.worksheet.max_row, 1, -1):  # 从最后一行开始删除
            self.worksheet.delete_rows(row)

        self.save()
        print("测试结果已清空")

    def get_test_results(self) -> List[Dict[str, Any]]:
        """
        获取所有测试结果
        :return: 测试结果列表
        """
        results = []
        headers = []

        # 读取表头
        for col in range(1, self.worksheet.max_column + 1):
            header = self.worksheet.cell(row=1, column=col).value
            if header:
                headers.append(header)

        # 读取数据
        for row in range(2, self.worksheet.max_row + 1):
            result_data = {}
            for col, header in enumerate(headers):
                value = self.worksheet.cell(row=row, column=col + 1).value
                result_data[header] = value

            if result_data:  # 跳过空行
                results.append(result_data)

        return results

    def close(self):
        """关闭Excel文件"""
        if self.workbook:
            self.workbook.close()

    def __del__(self):
        """析构函数，确保文件关闭"""
        self.close()


# 兼容旧版接口 - 工具函数
def read_pts_excel_config(excel_path: str, sheet_name: str = "测试配置") -> Dict[str, Any]:
    """
    读取PTS Excel配置文件
    :param excel_path: Excel文件路径
    :param sheet_name: 配置sheet名称
    :return: 配置字典
    """
    try:
        # 使用pandas读取更高效
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        # 转换为字典格式
        config_data = {}

        # 按测试类型分组
        for _, row in df.iterrows():
            test_type = row.get("测试类型", "default")
            if test_type not in config_data:
                config_data[test_type] = []

            config_data[test_type].append({
                "test_id": row.get("测试项编号"),
                "name": row.get("测试项名称"),
                "var_name": row.get("变量名"),
                "min": row.get("下限值"),
                "max": row.get("上限值"),
                "length": row.get("数据长度", 4),
                "switch_var": row.get("开关变量", "")
            })

        return config_data

    except Exception as e:
        raise Exception(f"读取Excel配置失败: {str(e)}")


def write_pts_test_result(excel_path: str, test_results: List[Dict[str, Any]], append: bool = True):
    """
    快速写入测试结果到PTS Excel
    :param excel_path: Excel文件路径
    :param test_results: 测试结果列表
    :param append: 是否追加写入（True=追加，False=覆盖）
    """
    try:
        if append and os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
            # 追加模式：使用已有文件
            manager = PTSExcelManager(excel_path, excel_path)
        else:
            # 覆盖模式：使用模板或创建新文件
            manager = PTSExcelManager(excel_path, excel_path)
            if append:  # 如果是追加模式但文件为空，需要初始化表头
                manager._init_header()

        manager.write_batch_results(test_results)
        manager.close()

    except Exception as e:
        raise Exception(f"写入测试结果失败: {str(e)}")


class ExcelHandler:
    """
    Excel文件操作封装类，基于openpyxl实现，支持单元格读写、类型查询、列值查找等功能
    已将sheet_name设为实例变量，调用方法时无需重复传入
    """

    def __init__(self, file_path, sheet_name):
        """
        初始化Excel处理器，加载指定路径的Excel文件和指定工作表
        :param file_path: Excel文件的完整路径
        :param sheet_name: 要操作的工作表名称（全局生效，无需重复传入）
        """
        self.file_path = file_path
        self.sheet_name = sheet_name  # 新增：将sheet_name设为实例变量（全局属性）
        self.workbook = None  # 工作簿对象，延迟加载或统一管理
        self.current_sheet = None  # 当前操作的工作表对象

    def _load_workbook(self, data_only=False):
        """
        私有方法：加载Excel工作簿（内部复用，避免重复加载）
        :param data_only: 是否只读取单元格计算结果（忽略公式），默认False
        """
        if not self.workbook:
            try:
                self.workbook = openpyxl.load_workbook(self.file_path, data_only=data_only)
            except FileNotFoundError:
                raise FileNotFoundError(f"错误：文件 '{self.file_path}' 不存在")
            except Exception as e:
                raise Exception(f"加载Excel文件失败：{e}")
        return self.workbook

    def _get_sheet(self, data_only=False):
        """
        私有方法：获取指定名称的工作表（内部复用，包含存在性校验）
        改动：不再接收sheet_name参数，直接使用实例变量self.sheet_name
        :param data_only: 是否只读取单元格计算结果
        :return: 工作表对象
        """
        wb = self._load_workbook(data_only=data_only)
        if self.sheet_name in wb.sheetnames:
            self.current_sheet = wb[self.sheet_name]
            return self.current_sheet
        else:
            raise ValueError(f"错误：工作表 '{self.sheet_name}' 不存在于文件 '{self.file_path}' 中")

    def get_cell_value(self, cell_address, data_only=False):
        """
        公开方法：获取指定单元格内容
        改动：移除sheet_name参数，直接使用实例变量
        :param cell_address: 单元格地址（如 'A1'）
        :param data_only: 是否只读取单元格计算结果（忽略公式）
        :return: 单元格的值
        """
        sheet = self._get_sheet(data_only=data_only)
        cell_value = sheet[cell_address].value
        return cell_value

    def get_cell_type(self, cell_address):
        """
        公开方法：获取指定单元格内容类型
        改动：移除sheet_name参数，直接使用实例变量
        :param cell_address: 单元格地址（如 'A1'）
        :return: 单元格数据类型（对应openpyxl的cell data_type）
        """
        sheet = self._get_sheet()
        cell_type = sheet[cell_address].data_type
        return cell_type

    def write_cell_value(self, cell_address, value):
        """
        公开方法：向指定单元格写入值
        改动：移除sheet_name参数，直接使用实例变量
        :param cell_address: 单元格地址（如 'A1'）
        :param value: 要写入的值
        """
        try:
            sheet = self._get_sheet()
            # 向指定的单元格写入值
            sheet[cell_address] = value
            # 保存修改
            self.workbook.save(self.file_path)
            # print(
            #     f"值 '{value}' 已写入文件 '{self.file_path}' 的工作表 '{self.sheet_name}' 中的单元格 '{cell_address}'")
        except Exception as e:
            raise Exception(f"写入单元格失败：{e}")

    def find_value_in_column(self, column_letter, value_to_find):
        """
        公开方法：在指定列中查找指定内容
        改动：移除sheet_name参数，直接使用实例变量
        :param column_letter: 列字母（如 'A'、'B'）
        :param value_to_find: 要查找的目标值
        :return: 所有匹配值的单元格位置列表（如 ['A5', 'A10']）
        """
        sheet = self._get_sheet()
        positions = []  # 用于保存找到值的位置
        column_index = column_index_from_string(column_letter)

        # 遍历指定列中的每个单元格
        for row in sheet.iter_rows(
                min_col=column_index,
                max_col=column_index,
                values_only=False
        ):
            for cell in row:
                if cell.value == value_to_find:
                    # 保存找到的单元格位置 (如 A5)
                    positions.append(cell.coordinate)
                    print('行号是', cell.row)

        return positions

    def find_value_in_cell_return_row(self, column_number, value_to_find):
        """
        公开方法：根据列号和值查找行号
        :param column_number: 列号
        :param value_to_find: 要查找的值
        :return: 行号
        """
        sheet = self._get_sheet()
        row_number = sheet.max_row
        column_index = column_index_from_string(column_number)

        # 遍历指定列中的每个单元格
        for row in sheet.iter_rows(
                min_col=column_index,
                max_col=column_index,
                values_only=False
        ):
            for cell in row:
                if cell.value == value_to_find or value_to_find in cell.value:
                    # 保存找到的单元格位置 (如 A5)
                    # print('行号是', cell.row)
                    return cell.row

    def get_row_from_cell(self, cell_position, data_only=True):
        """
        公开方法：根据单元格位置返回其所在行号
        改动：移除sheet_name参数，直接使用实例变量
        :param cell_position: 单元格位置（例如 "B5"）
        :param data_only: 是否只读取单元格计算结果
        :return: 单元格所在的行号
        """
        sheet = self._get_sheet(data_only=data_only)
        # 获取单元格对象并返回行号
        cell = sheet[cell_position]
        return cell.row

    def update_sheet_name(self, new_sheet_name):
        """
        新增公开方法：更新实例的工作表名称（支持切换工作表）
        :param new_sheet_name: 新的工作表名称
        """
        # 先校验新工作表是否存在
        wb = self._load_workbook()
        if new_sheet_name in wb.sheetnames:
            self.sheet_name = new_sheet_name
            self.current_sheet = wb[new_sheet_name]
            print(f"工作表已切换为：'{new_sheet_name}'")
        else:
            raise ValueError(f"错误：工作表 '{new_sheet_name}' 不存在于文件 '{self.file_path}' 中")

    def close_workbook(self):
        """
        公开方法：手动关闭工作簿，释放资源
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.current_sheet = None
            print(f"Excel文件 '{self.file_path}' 已关闭，资源已释放")

    def __del__(self):
        """
        析构方法：对象销毁时自动关闭工作簿，防止资源泄露
        """
        self.close_workbook()


# 使用示例
if __name__ == "__main__":
    # 示例1: 创建或追加测试结果
    template_path = "template.xlsx"  # 模板文件
    output_path = "test_results.xlsx"  # 输出文件

    # 测试数据
    test_data = {
        "test_id": "TC001",
        "test_name": "电压测试",
        "var_name": "Voltage",
        "low_limit": 3.0,
        "up_limit": 3.6,
        "actual_value": 3.45,
        "test_result": "Pass",
        "remarks": "首次测试"
    }

    # 创建管理器
    manager = PTSExcelManager(template_path, output_path)

    # 写入测试结果（会追加到已有数据后面）
    manager.write_test_result(test_data)

    # 批量写入
    more_results = [
        {
            "test_id": "TC002",
            "test_name": "电流测试",
            "var_name": "Current",
            "low_limit": 100,
            "up_limit": 200,
            "actual_value": 150,
            "test_result": "Pass",
            "remarks": ""
        }
    ]
    manager.write_batch_results(more_results)

    # 读取所有测试结果
    all_results = manager.get_test_results()
    print(f"共{len(all_results)}条测试结果")

    # 关闭管理器
    manager.close()

    # 使用快捷函数
    quick_results = [
        {
            "test_id": "TC003",
            "test_name": "温度测试",
            "var_name": "Temperature",
            "low_limit": -20,
            "up_limit": 80,
            "actual_value": 25,
            "test_result": "Pass",
            "remarks": "环境温度"
        }
    ]

    # 追加写入（不会覆盖已有数据）
    write_pts_test_result(output_path, quick_results, append=True)

    print("测试结果写入完成")
